import sys
import os

import openpyxl

sys.path.insert(0, './whatsApp')
import whatsApp.mainwindow as MainWindow
from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableWidgetItem
from PySide6.QtGui import QIcon, QPixmap
from configparser import ConfigParser

config = ConfigParser()
config.read('whatsapp_config.ini')
config.sections()
company_name = config["comp_name"]["name"]
company_cont = config["comp_name"]["number"]


class MainWindow(QMainWindow, MainWindow.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon('./whatsApp/wpicon.png'))
        self.setWindowTitle('WhatsAppBot')
        self.editTable()
        self.add_excel_button.clicked.connect(self.add_file)
        self.add_excel_button_3.clicked.connect(self.clear_list)
        self.selectAllButton.clicked.connect(self.sel_all)
        self.sendButton.clicked.connect(self.send_message)

    def sel_all(self):
        self.tableWidget.selectAll()

    def clear_list(self):
        self.tableWidget.clearContents()

    def editTable(self):
        self.tableWidget.setColumnWidth(2, 550)

    def add_file(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            caption='Add Songs to the app',
            dir=':\\',
            filter='Supported Files (*.xlsm;*.xlsx;*txt)'
        )
        if files:
            workbook = openpyxl.load_workbook(files[0])
            sheet = workbook.active
            self.tableWidget.setRowCount(sheet.max_row + 1)
            self.tableWidget.setColumnCount(sheet.max_column)
            list_values = list(sheet.values)
            #self.tableWidget.setHorizontalHeaderLabels(list_values[0])
            row_index = 0
            for value_tuple in list_values[0:]:
                col_index = 0
                for value in value_tuple:
                    self.tableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                    col_index += 1
                row_index += 1
    # data 0 : içeriği elle gir aktif, data 1 : bakiye mesaj içeriği, data 2 :  bakiye hatırlatma, data 3: bayram içerik
    def message_builder(self, name,data) -> str:
        if data[0] == "otomatik":
            return data[1]
        if data[0] == "True":

            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak sizlere iyi günler dileriz."
                             f"Hesaplarımızda {data[1]} 'TL tutarında bir bakiyeniz görülmektedir"
                             f"Mevcut bakiyenizi en kısa sürede ödemenizi rica ederiz."
                             f"Bu mesaj otomasyon tarafından gönderilmiştir."
                             f"Bir hata olduğunu düşünüyorsanız lütfen bu numara üzerinden irtibata geçiniz")
            return message_total
        if data[0]  == "Kurban Bayramı":

            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak Kurban Bayramınızı"
                             f"en içten dileklerimiz ile kutlar, size ve ailenize sağlık, mutluluk, esenlikler dileriz")
            return message_total
        if data[0]  == "Ramazan Bayramı":

            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak Ramazan Bayramınızı"
                             f"en içten dileklerimiz ile kutlar, size ve ailenize sağlık, mutluluk, esenlikler dileriz")
            return message_total
        if data[0]  == "Kandil":

            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak {data[1]} kandilinizi"
                             f"en içten dileklerimiz ile kutlar, hayırlara vesile olmasını dileriz.")
            return message_total
        if data[0]  == "Bayram":

            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak {data[1]} bayramınızı"
                             f"en içten dileklerimiz ile kutlar, Cumhuriyetimizin kurucusu Gazi Mustafa Kemal Atatürk başta olmak üzere silah arkadaşlarını ve tüm şehitlerimizi rahmetle, kahraman gazilerimizi minnet ve şükranla anıyoruz;.")
            return message_total


    def send_message(self):
        content = {}
        paths = []
        data = []
        if self.checkBox.isChecked():
            choice =  self.radioButton_2.isChecked() or self.comboBox.currentText()
        else:
            choice = "otomatik"
        choice = str(choice)
        data[0] = choice

        selected = self.tableWidget.selectedItems()
        if selected:
            for item in selected:
                if not item.row() in paths:
                    paths.append(item.row())
        if len(paths) > 0:
            for path in paths:
                data_isim = self.tableWidget.item(path, 0).text()
                data_numara = self.tableWidget.item(path, 1).text()
                #self.message_builder(data_isim, [])
                data_mesaj = self.tableWidget.item(path, 2).text()

                content[data_numara] = data_mesaj
        #print(content)
        #self.message_builder(company_name, content)
        #from whatsapp_modul import *


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
